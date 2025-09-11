"""
SLA Report Processing Module

This module handles the complete processing of SLA data files including:
- Excel/CSV parsing
- Date/time calculations  
- SLA business logic application
- Report generation with proper formatting

Moved from frontend to enable server-side processing.
"""

import os
import re
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
from typing import List, Dict, Any, Tuple, Optional
from io import BytesIO
import json
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# Constants
YELLOW_FIELDS = [
    "ResolSLA", "RespSLA", "ReqComp", "ReqCrDtConc", "EnDtConc", 
    "HisChDtTiConc", "ElapsedTime", "CalcPreDt", "RefinedPreDt", 
    "CalcStDt", "RefinedStDt", "Cumilative", "ResolSOW", "RespSOW", 
    "ResolRem", "RespRem", "Rollover", "ReqCrYM", "DateRollover", "DateReqCrYM"
]

# Holiday data for multiple years (2021-2025)
HOLIDAYS_BY_YEAR = {
    "2020": [
        "2020-01-01", "2020-01-14", "2020-01-26", "2020-04-10", "2020-04-13",
        "2020-04-21", "2020-05-01", "2020-08-15", "2020-08-27", "2020-10-02",
        "2020-11-20", "2020-12-25"
    ],
    "2021": [
        "2021-01-01", "2021-01-14", "2021-02-15", "2021-04-02", "2021-04-05",
        "2021-04-21", "2021-05-01", "2021-08-15", "2021-08-27", "2021-10-02",
        "2021-11-11", "2021-12-25"
    ],
    "2022": [
        "2022-01-01", "2022-01-14", "2022-03-07", "2022-04-15", "2022-04-18",
        "2022-04-21", "2022-05-01", "2022-08-15", "2022-08-27", "2022-10-02",
        "2022-11-02", "2022-12-25"
    ],
    "2023": [
        "2023-01-01", "2023-01-14", "2023-02-20", "2023-04-07", "2023-04-10",
        "2023-04-21", "2023-05-01", "2023-08-15", "2023-08-27", "2023-10-02",
        "2023-11-13", "2023-12-25"
    ],
    "2024": [
        "2024-01-01", "2024-01-14", "2024-01-26", "2024-03-29", "2024-04-02",
        "2024-04-13", "2024-04-14", "2024-04-21", "2024-06-02", "2024-08-15",
        "2024-09-10", "2024-10-02", "2024-10-31", "2024-12-25"
    ],
    "2025": [
        "2025-01-01", "2025-01-14", "2025-02-27", "2025-03-31", "2025-05-01",
        "2025-08-15", "2025-08-27", "2025-10-02", "2025-10-21", "2025-12-25"
    ]
}

SLA_TABLE = {
    "P1": {"respsow": 0.5, "resolsow": 4},
    "P2": {"respsow": 2, "resolsow": 9},
    "P3": {"respsow": 9, "resolsow": 45},
    "P4": {"respsow": 18, "resolsow": 90},
}

WORK_HOURS = {
    "start": "14:00:00",
    "end": "23:00:00",
}

REPORT_PREFIX = "EI_Inc_Powbi"

class DateUtils:
    """Date and time utility functions"""
    
    @staticmethod
    def excel_serial_to_date(serial):
        """Convert Excel serial number to date - MATCHES FRONTEND LOGIC EXACTLY"""
        # Frontend: if (!serial || isNaN(serial)) return null;
        if not serial or pd.isna(serial):
            return None
        try:
            if isinstance(serial, (int, float)):
                # Frontend logic: const excelEpoch = new Date(1900, 0, 1); const days = Math.floor(serial) - 2; return new Date(excelEpoch.getTime() + days * 24 * 60 * 60 * 1000);
                # But since frontend uses the other method for actual conversion, let's use the same approach
                excel_serial = float(serial)
                days_since_unix_epoch = excel_serial - 25569
                milliseconds = round(days_since_unix_epoch * 86400 * 1000)
                return datetime.utcfromtimestamp(milliseconds / 1000.0)
            return None
        except:
            return None
    
    @staticmethod
    def format_time(time_str):
        """Format time string to HH:MM:SS - matches frontend logic exactly"""
        # Frontend: if (!timeStr || typeof timeStr !== "string") return "00:00:00";
        if not time_str or pd.isna(time_str):
            return "00:00:00"
        
        # Frontend: timeStr = timeStr.replace(/[^0-9]/g, "").padStart(6, "0");
        # Convert to string and remove all non-numeric characters (same as frontend)
        time_str = re.sub(r'[^0-9]', '', str(time_str)).zfill(6)
        
        # Frontend: return timeStr.length >= 6 ? `${timeStr.slice(0, 2)}:${timeStr.slice(2, 4)}:${timeStr.slice(4, 6)}` : "00:00:00";
        if len(time_str) >= 6:
            return f"{time_str[:2]}:{time_str[2:4]}:{time_str[4:6]}"
        return "00:00:00"
    
    @staticmethod
    def format_date(date):
        """Format date to MM/DD/YYYY"""
        if date is None or pd.isna(date):
            return ""
        if not hasattr(date, 'month') or not hasattr(date, 'day') or not hasattr(date, 'year'):
            return ""
        try:
            return f"{date.month:02d}/{date.day:02d}/{date.year}"
        except:
            return ""
    
    @staticmethod
    def parse_date_time(date_val, time_str, index=None):
        """Parse date and time values into datetime object - MATCHES FRONTEND LOGIC EXACTLY"""
        try:
            # Frontend: if (!dateVal || dateVal === null) return null;
            if date_val is None or pd.isna(date_val):
                return None
                
            date = None
            
            # Frontend: if (typeof dateVal === "number") { date = dateUtils.excelSerialToDate(dateVal); }
            if isinstance(date_val, (int, float)) and not pd.isna(date_val):
                # Use frontend-compatible excel serial conversion
                try:
                    excel_serial = float(date_val)
                    days_since_unix_epoch = excel_serial - 25569
                    milliseconds = round(days_since_unix_epoch * 86400 * 1000)
                    date = datetime.utcfromtimestamp(milliseconds / 1000.0)
                except:
                    return None
                    
            # Frontend: else if (typeof dateVal === "string") { const datePart = dateVal?.split(" ")[0]; const [day, month, year] = datePart.split('/'); const swappedDate = `${month}/${day}/${year}`; date = new Date(swappedDate); }
            elif isinstance(date_val, str) and date_val.strip():
                # Handle ISO datetime format: YYYY-MM-DD HH:MM:SS or YYYY-MM-DD
                if '-' in date_val and len(date_val) >= 10:
                    date_part = date_val.split(" ")[0] if " " in date_val else date_val
                    if len(date_part) >= 10:  # YYYY-MM-DD format
                        try:
                            parts = date_part.split('-')
                            if len(parts) == 3:
                                year, month, day = parts
                                date = datetime(int(year), int(month), int(day))
                        except (ValueError, TypeError):
                            return None
                # Handle slash-separated dates: DD/MM/YYYY or MM/DD/YYYY
                elif '/' in date_val:
                    date_part = date_val.split(" ")[0] if " " in date_val else date_val
                    parts = date_part.split('/')
                    if len(parts) == 3:
                        a, b, c = parts
                        # Detect DD/MM/YYYY vs MM/DD/YYYY format
                        try:
                            a_int, b_int = int(a), int(b)
                            
                            # If first part > 12, it must be DD/MM/YYYY format
                            if a_int > 12:
                                day, month, year = a, b, c
                            # If second part > 12, it must be MM/DD/YYYY format
                            elif b_int > 12:
                                month, day, year = a, b, c
                            else:
                                # Both <= 12: ambiguous, try MM/DD/YYYY first then DD/MM/YYYY
                                month, day, year = a, b, c
                            
                            # Create date object using detected format
                            date = datetime(int(year), int(month), int(day))
                        except (ValueError, TypeError):
                            # If MM/DD/YYYY failed and both parts <= 12, try DD/MM/YYYY
                            try:
                                if a_int <= 12 and b_int <= 12:
                                    day, month, year = a, b, c
                                    date = datetime(int(year), int(month), int(day))
                                else:
                                    return None
                            except (ValueError, TypeError):
                                return None
                            
            # Frontend: else { date = new Date(dateVal); }
            elif date_val and not pd.isna(date_val):
                try:
                    date = pd.to_datetime(date_val, errors='coerce')
                    if pd.isna(date):
                        return None
                except:
                    return None
                
            # Frontend: if (!date || isNaN(date)) return null;
            if not date or pd.isna(date):
                return None
                
            # Frontend: const time = dateUtils.formatTime(timeStr); return new Date(`${dateUtils.formatDate(date)} ${time}`);
            time_formatted = DateUtils.format_time(time_str)
            
            # Frontend formatDate returns MM/DD/YYYY format
            date_formatted = DateUtils.format_date(date)  # This returns MM/DD/YYYY
            
            # Create combined datetime string and parse it
            try:
                combined_str = f"{date_formatted} {time_formatted}"
                result_date = datetime.strptime(combined_str, "%m/%d/%Y %H:%M:%S")
                return result_date
            except (ValueError, AttributeError):
                # Fallback to date without time if time parsing fails
                return date
                
        except Exception as e:
            print(f"⚠️ Warning: Error parsing date/time at index {index}: {e}")
            return None
    
    @staticmethod
    def to_mmddyyyy_from_any(date_val, parsed_datetime=None):
        """Normalize many date representations to MM/DD/YYYY string for output and further parsing."""
        try:
            # If we already have a parsed datetime, format it
            if parsed_datetime is not None and not pd.isna(parsed_datetime):
                return DateUtils.format_date(parsed_datetime)
            
            if date_val is None or pd.isna(date_val):
                return ""
            
            # If it's a pandas/py datetime-like
            if hasattr(date_val, 'year') and hasattr(date_val, 'month') and hasattr(date_val, 'day'):
                try:
                    return f"{date_val.month:02d}/{date_val.day:02d}/{date_val.year}"
                except:
                    pass
            
            # If it's a string
            s = str(date_val).strip()
            if not s:
                return ""
            
            # Handle ISO-like: YYYY-MM-DD or with time
            if '-' in s and len(s) >= 10:
                date_part = s.split(' ')[0]
                parts = date_part.split('-')
                if len(parts) == 3 and len(parts[0]) == 4:
                    year, month, day = parts
                    return f"{month.zfill(2)}/{day.zfill(2)}/{year}"
            
            # Handle slash dates - IMPROVED LOGIC FOR DD/MM/YYYY DETECTION
            if '/' in s:
                parts = s.split(' ')[0].split('/')
                if len(parts) == 3:
                    a, b, c = parts
                    # c should be year (4 digits)
                    if len(c) == 4 and a.isdigit() and b.isdigit():
                        a_int, b_int = int(a), int(b)
                        
                        # If first part > 12, it must be day (DD/MM/YYYY format)
                        if a_int > 12:
                            day, month, year = a, b, c
                            return f"{str(month).zfill(2)}/{str(day).zfill(2)}/{year}"
                        # If second part > 12, format is likely MM/DD/YYYY already  
                        elif b_int > 12:
                            month, day, year = a, b, c
                            return f"{str(month).zfill(2)}/{str(day).zfill(2)}/{year}"
                        # Both parts <= 12: ambiguous, but assume DD/MM/YYYY based on input data pattern
                        else:
                            day, month, year = a, b, c
                            return f"{str(month).zfill(2)}/{str(day).zfill(2)}/{year}"
            
            # Try excel serial
            if isinstance(date_val, (int, float)) and not pd.isna(date_val):
                dt = DateUtils.excel_serial_to_date(date_val)
                if dt:
                    return f"{dt.month:02d}/{dt.day:02d}/{dt.year}"
        except Exception:
            return ""
        return ""

    @staticmethod
    def convert_excel_date(excel_date):
        """Convert Excel date to DD/MM/YYYY format - MATCHES FRONTEND EXACTLY"""
        # Handle None, NaN, NaT values first
        if excel_date is None or pd.isna(excel_date):
            return ""
        
        # Handle string dates that already contain "/" - CONVERT DD/MM/YYYY to DD/MM/YYYY (keep same format for now)
        if isinstance(excel_date, str) and "/" in excel_date:
            # Input is typically DD/MM/YYYY format, return as-is for now
            # The conversion to MM/DD/YYYY happens later in to_mmddyyyy_from_any
            return excel_date
            
        # Handle numeric Excel serial dates - MATCH FRONTEND LOGIC EXACTLY
        if isinstance(excel_date, (int, float)) and not pd.isna(excel_date):
            try:
                # Frontend logic: const date = new Date(Math.round((excelDate - 25569) * 86400 * 1000));
                # Convert to milliseconds since Unix epoch, then create datetime
                excel_serial = float(excel_date)
                # Excel epoch is 25569 days before Unix epoch (Jan 1, 1970)
                days_since_unix_epoch = excel_serial - 25569
                milliseconds = round(days_since_unix_epoch * 86400 * 1000)
                
                # Create datetime from milliseconds
                date = datetime.utcfromtimestamp(milliseconds / 1000.0)
                
                # Return in DD/MM/YYYY format (en-GB locale equivalent)
                return date.strftime("%d/%m/%Y")
            except Exception as e:
                print(f"⚠️ Warning: Excel date conversion failed for {excel_date}: {e}")
                return str(excel_date)
                
        # Handle datetime objects (but check for NaT first)
        if hasattr(excel_date, 'strftime') and not pd.isna(excel_date):
            try:
                return excel_date.strftime("%d/%m/%Y")
            except:
                return str(excel_date)
                
        # Fallback for any other type
        return str(excel_date) if excel_date else ""

class CalculationUtils:
    """Business calculation utilities"""
    
    @staticmethod
    def network_days_intl(start_date, end_date, holidays=None):
        """Calculate working days between dates excluding weekends and holidays"""
        if not start_date or not end_date or pd.isna(start_date) or pd.isna(end_date):
            return 0
            
        holidays = holidays or []
        holiday_dates = [pd.to_datetime(h).normalize() for h in holidays]
        
        days = 0
        current = pd.to_datetime(start_date).normalize()
        end = pd.to_datetime(end_date).normalize()
        
        while current <= end:
            if current.weekday() < 5 and current not in holiday_dates:  # Monday=0, Sunday=6
                days += 1
            current += timedelta(days=1)
        
        return days
    
    @staticmethod
    def is_working_day(date, holidays=None):
        """Check if a date is a working day"""
        if not date or pd.isna(date):
            return False
        
        holidays = holidays or []
        date_normalized = pd.to_datetime(date).normalize()
        holiday_dates = [pd.to_datetime(h).normalize() for h in holidays]
        
        # Check if it's weekend (Saturday=5, Sunday=6)
        if date_normalized.weekday() >= 5:
            return False
        
        # Check if it's a holiday
        if date_normalized in holiday_dates:
            return False
            
        return True
    
    @staticmethod 
    def calculate_pre_dt(start_date, end_date, work_start_time_str, work_end_time_str, holidays=None, index=None):
        """Calculate working hours between two dates - MATCHES FRONTEND calculatePreDt EXACTLY"""
        # Frontend: if (!startDate || !endDate || isNaN(startDate) || isNaN(endDate)) return 0;
        if not start_date or not end_date or pd.isna(start_date) or pd.isna(end_date):
            return 0
        
        try:
            start_dt = pd.to_datetime(start_date) if not hasattr(start_date, 'hour') else start_date
            end_dt = pd.to_datetime(end_date) if not hasattr(end_date, 'hour') else end_date
        except:
            return 0
        
        if end_dt < start_dt:
            return 0
            
        # Frontend: const [workStartHours, workStartMinutes] = workStartTimeStr.split(":").map(Number);
        work_start_parts = work_start_time_str.split(":")
        work_end_parts = work_end_time_str.split(":")
        
        work_start_hours = int(work_start_parts[0])
        work_start_minutes = int(work_start_parts[1])
        work_end_hours = int(work_end_parts[0])
        work_end_minutes = int(work_end_parts[1])
        
        # Frontend: const workDayStart = workStartHours + workStartMinutes / 60;
        work_day_start = work_start_hours + work_start_minutes / 60
        work_day_end = work_end_hours + work_end_minutes / 60
        work_day_length = work_day_end - work_day_start
        
        holidays = holidays or []
        
        # Frontend: const holidayTimestamps = holidays.map((h) => new Date(h).setHours(0, 0, 0, 0));
        holiday_timestamps = [pd.to_datetime(h).normalize() for h in holidays]
        
        # Frontend: const isWorkingDay = (date) => {...}
        def is_working_day(date):
            day = date.weekday()  # Monday=0, Sunday=6 in pandas (vs 0=Sunday in JS)
            # Convert to JS weekday: Monday=1, Sunday=0
            js_day = (day + 1) % 7
            date_timestamp = pd.to_datetime(date).normalize()
            return js_day != 0 and js_day != 6 and date_timestamp not in holiday_timestamps
        
        # Frontend: let adjustedStartDate = new Date(startDate);
        adjusted_start_date = pd.to_datetime(start_dt)
        
        # Frontend: if (!isWorkingDay(adjustedStartDate)) { while (!isWorkingDay(adjustedStartDate)) {...} }
        if not is_working_day(adjusted_start_date):
            while not is_working_day(adjusted_start_date):
                adjusted_start_date = adjusted_start_date + timedelta(days=1)
                adjusted_start_date = adjusted_start_date.normalize()
            # Frontend: adjustedStartDate.setHours(workStartHours, workStartMinutes, 0, 0);
            adjusted_start_date = adjusted_start_date.replace(hour=work_start_hours, minute=work_start_minutes, second=0)
        
        # Frontend: const networkDays = calculationUtils.networkDaysIntl(adjustedStartDate, endDate, holidays);
        network_days = CalculationUtils.network_days_intl(adjusted_start_date, end_dt, holidays)
        
        # Frontend: if (networkDays === 0) return 0;
        if network_days == 0:
            return 0
        
        # Frontend: const getMedTime = (date) => {...}
        def get_med_time(date):
            if not is_working_day(date):
                return work_day_end
            # Frontend: const hours = date.getHours() + date.getMinutes() / 60 + date.getSeconds() / 3600;
            hours = date.hour + date.minute / 60 + date.second / 3600
            # Frontend: return Math.max(workDayStart, Math.min(hours, workDayEnd));
            return max(work_day_start, min(hours, work_day_end))
        
        # Frontend: const startMedTime = getMedTime(adjustedStartDate);
        start_med_time = get_med_time(adjusted_start_date)
        end_med_time = get_med_time(end_dt)
        
        # Frontend calculation logic
        result = 0
        if network_days == 1:
            # Frontend: if (isWorkingDay(adjustedStartDate) && isWorkingDay(endDate)) { result = endMedTime - startMedTime; }
            if is_working_day(adjusted_start_date) and is_working_day(end_dt):
                result = end_med_time - start_med_time
        else:
            # Frontend: const fullDaysPart = (networkDays - 1) * workDayLength; result = fullDaysPart + (endMedTime - startMedTime);
            full_days_part = (network_days - 1) * work_day_length
            result = full_days_part + (end_med_time - start_med_time)
        
        # Frontend: result = result > 0 ? parseFloat(result.toFixed(2)) : 0.0;
        result = float(f"{result:.2f}") if result > 0 else 0.0
        return result

def parse_custom_date(date_string):
    """Parse custom date string format - MATCHES FRONTEND LOGIC EXACTLY"""
    # Frontend: const [datePart, timePart] = dateString.split(' '); const [month, day, year] = datePart.split('/');
    if not date_string or pd.isna(date_string):
        return None
    
    try:
        date_string = str(date_string).strip()
        if not date_string or date_string == " ":
            return None
            
        parts = date_string.split(' ')
        if len(parts) >= 2:
            date_part, time_part = parts[0], parts[1]
            date_parts = date_part.split('/')
            time_parts = time_part.split(':')
            
            if len(date_parts) == 3 and len(time_parts) == 3:
                try:
                    # Frontend logic: const [month, day, year] = datePart.split('/');
                    # The frontend expects MM/DD/YYYY format in custom date strings
                    month, day, year = [int(x) for x in date_parts]
                    hours, minutes, seconds = [int(x) for x in time_parts]
                    
                    # Frontend: return new Date(`${year}-${month}-${day}T${hours}:${minutes}:${seconds}`);
                    # Validate date components
                    if 1 <= month <= 12 and 1 <= day <= 31 and year > 1900:
                        return datetime(year, month, day, hours, minutes, seconds)
                    else:
                        return None
                except (ValueError, TypeError):
                    # If parsing fails, try without time
                    try:
                        month, day, year = [int(x) for x in date_parts]
                        if 1 <= month <= 12 and 1 <= day <= 31 and year > 1900:
                            return datetime(year, month, day)
                        else:
                            return None
                    except (ValueError, TypeError):
                        return None
    except Exception as e:
        print(f"⚠️ Warning: Error parsing custom date '{date_string}': {e}")
        return None
    
    return None

def convert_to_iso_date(date_str):
    """Convert DD/MM/YYYY to YYYY-MM-DD format"""
    if not date_str:
        return ''
    
    try:
        parts = date_str.split('/')
        if len(parts) == 3:
            a, b, c = parts
            # Assume DD/MM/YYYY format (based on convert_excel_date output)
            if len(c) == 4:  # Year is in position 3
                day, month, year = a, b, c
                if len(day) <= 2 and len(month) <= 2:
                    return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    except:
        pass
    
    return ''

def get_holidays_for_years(years):
    """Get holidays for specific years"""
    unique_years = list(set(str(year) for year in years if year))
    all_holidays = []
    
    for year in unique_years:
        if year in HOLIDAYS_BY_YEAR:
            all_holidays.extend(HOLIDAYS_BY_YEAR[year])
    
    return all_holidays

def extract_date_time_from_filename(name):
    """Extract date-time substring from filename"""
    if not name or not isinstance(name, str):
        return None
    
    base = os.path.splitext(name)[0]
    
    # Various date-time patterns
    patterns = [
        r'(\d{2})[-_\/]?(\d{2})[-_\/]?(\d{4})[\s_-]?(\d{2})[-_:]?(\d{2})[-_:]?(\d{2})',  # DD-MM-YYYY HH:MM:SS
        r'(\d{4})[-_\/]?(\d{2})[-_\/]?(\d{2})[\s_-]?(\d{2})[-_:]?(\d{2})[-_:]?(\d{2})',  # YYYY-MM-DD HH:MM:SS
        r'(\d{2})(\d{2})(\d{4})[\s_-]?(\d{2})[-_:]?(\d{2})[-_:]?(\d{2})',  # DDMMYYYY HHMMSS
        r'(\d{4})(\d{2})(\d{2})[\s_-]?(\d{2})[-_:]?(\d{2})[-_:]?(\d{2})',  # YYYYMMDD HHMMSS
    ]
    
    for pattern in patterns:
        match = re.search(pattern, base)
        if match:
            groups = match.groups()
            if len(groups) == 6:
                if len(groups[0]) == 4:  # YYYY format
                    yyyy, mm, dd, hh, mi, ss = groups
                else:  # DD format
                    dd, mm, yyyy, hh, mi, ss = groups
                return f"{yyyy}{mm}{dd}_{hh}{mi}{ss}"
    
    # Date-only patterns
    date_patterns = [
        r'(\d{2})[-_\/]?(\d{2})[-_\/]?(\d{4})',  # DD/MM/YYYY
        r'(\d{4})[-_\/]?(\d{2})[-_\/]?(\d{2})',  # YYYY/MM/DD
        r'(\d{2})(\d{2})(\d{4})',  # DDMMYYYY
        r'(\d{4})(\d{2})(\d{2})',  # YYYYMMDD
    ]
    
    for pattern in date_patterns:
        match = re.search(pattern, base)
        if match:
            groups = match.groups()
            if len(groups) == 3:
                if len(groups[0]) == 4:  # YYYY format
                    yyyy, mm, dd = groups
                else:  # DD format
                    dd, mm, yyyy = groups
                return f"{yyyy}{mm}{dd}_000000"
    
    return None

def generate_report_filename(source_name):
    """Generate output filename based on source filename"""
    extracted = extract_date_time_from_filename(source_name or '')
    if extracted:
        return f"{REPORT_PREFIX}_{extracted}.xlsx"
    
    now = datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    return f"{REPORT_PREFIX}_{timestamp}.xlsx"

class SLAProcessor:
    """Main SLA processing class"""
    
    def __init__(self):
        self.holidays = []
    
    def process_file_data(self, file_bytes: bytes, filename: str) -> Tuple[str, bytes]:
        """
        Process SLA data file and return processed Excel bytes
        
        Args:
            file_bytes: Raw file bytes (CSV or Excel)
            filename: Original filename
            
        Returns:
            Tuple of (output_filename, processed_file_bytes)
        """
        print(f"🔄 Processing SLA data file: {filename}")
        
        try:
            # Determine file type and parse
            file_extension = filename.lower().split('.')[-1]
            
            if file_extension == 'csv':
                # Read without any automatic date parsing to match frontend behavior exactly
                df = pd.read_csv(BytesIO(file_bytes), dtype=str, keep_default_na=False, na_values=[])
                # Replace any remaining NaN values with empty strings
                df = df.fillna('')
                data = [df.columns.tolist()] + df.values.tolist()
            elif file_extension in ['xlsx', 'xls']:
                # Read without any automatic date parsing to match frontend behavior exactly
                df = pd.read_excel(BytesIO(file_bytes), engine='openpyxl', dtype=str, keep_default_na=False, na_values=[])
                # Replace any remaining NaN values with empty strings
                df = df.fillna('')
                data = [df.columns.tolist()] + df.values.tolist()
            else:
                raise ValueError(f"Unsupported file type: {file_extension}")
            
            # Process the data
            processed_data = self._process_excel_data(data)
            
            # Generate output filename
            output_filename = generate_report_filename(filename)
            
            # Convert to Excel bytes
            output_bytes = self._create_excel_output(processed_data)
            
            print(f"✅ Processing completed: {output_filename}")
            return output_filename, output_bytes
            
        except Exception as e:
            print(f"❌ Error processing SLA file: {e}")
            print(f"📋 Error details: {type(e).__name__}: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
    
    def process_file_data_enhanced(
        self, 
        file_bytes: bytes, 
        filename: str, 
        output_format: str = "excel",
        clean_format: bool = True
    ) -> Tuple[str, bytes]:
        """
        Enhanced processing with format options and cleaner output
        
        Args:
            file_bytes: Raw file bytes (CSV or Excel)
            filename: Original filename
            output_format: "excel" or "csv"
            clean_format: If True, truncates long text fields for better visibility
            
        Returns:
            Tuple of (output_filename, processed_file_bytes)
        """
        print(f"🔄 Processing SLA data file with enhanced options: {filename}")
        
        try:
            # Determine file type and parse
            file_extension = filename.lower().split('.')[-1]
            
            if file_extension == 'csv':
                # Read without any automatic date parsing to match frontend behavior exactly
                df = pd.read_csv(BytesIO(file_bytes), dtype=str, keep_default_na=False, na_values=[])
                # Replace any remaining NaN values with empty strings
                df = df.fillna('')
                data = [df.columns.tolist()] + df.values.tolist()
            elif file_extension in ['xlsx', 'xls']:
                # Read without any automatic date parsing to match frontend behavior exactly
                df = pd.read_excel(BytesIO(file_bytes), engine='openpyxl', dtype=str, keep_default_na=False, na_values=[])
                # Replace any remaining NaN values with empty strings
                df = df.fillna('')
                data = [df.columns.tolist()] + df.values.tolist()
            else:
                raise ValueError(f"Unsupported file type: {file_extension}")
            
            # Process the data (same as before)
            processed_data = self._process_excel_data(data)
            
            # Clean the data if requested
            if clean_format:
                print(f"🧹 Applying visibility cleaning (truncating long text fields)")
                processed_data = self._clean_for_visibility(processed_data)
            else:
                print(f"📝 Skipping visibility cleaning - preserving full text content")
            
            # Generate output filename with correct extension
            base_name = generate_report_filename(filename)
            if output_format.lower() == "csv":
                output_filename = base_name.replace('.xlsx', '.csv')
            else:
                output_filename = base_name
            
            # Convert to requested format
            if output_format.lower() == "csv":
                output_bytes = self._create_csv_output(processed_data)
            else:
                output_bytes = self._create_excel_output(processed_data)
            
            print(f"✅ Enhanced processing completed: {output_filename}")
            return output_filename, output_bytes
            
        except Exception as e:
            print(f"❌ Error in enhanced processing: {e}")
            print(f"📋 Error details: {type(e).__name__}: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
    
    def _clean_for_visibility(self, processed_data):
        """Clean data for better visibility by truncating long text fields"""
        headers = processed_data[0]
        rows = processed_data[1:]
        
        # Find text columns that are typically very long
        text_columns = []
        for i, header in enumerate(headers):
            if any(text_field in str(header).lower() for text_field in 
                   ['text request', 'text answer', 'subject description']):
                text_columns.append(i)
        
        # Truncate long text fields
        cleaned_rows = []
        for row in rows:
            new_row = row[:]
            for col_idx in text_columns:
                if col_idx < len(new_row) and new_row[col_idx]:
                    text_val = str(new_row[col_idx])
                    if len(text_val) > 100:  # Truncate if longer than 100 chars
                        new_row[col_idx] = text_val[:97] + "..."
            cleaned_rows.append(new_row)
        
        return [headers] + cleaned_rows
    
    def _create_csv_output(self, processed_data):
        """Create CSV output from processed data"""
        headers = processed_data[0]
        rows = processed_data[1:]
        
        # Create DataFrame
        df = pd.DataFrame(rows, columns=headers)
        
        # Format numeric columns properly
        sla_columns = [col for col in df.columns if col in YELLOW_FIELDS]
        for col in sla_columns:
            if col in df.columns:
                try:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                except:
                    pass
        
        # Convert to CSV bytes
        output = BytesIO()
        df.to_csv(output, index=False, encoding='utf-8')
        output.seek(0)
        return output.getvalue()
    
    def _process_excel_data(self, data):
        """Process Excel data with SLA calculations"""
        if not data or len(data) == 0:
            raise ValueError("No data to process")
        
        headers = data[0]
        rows = data[1:]
        
        print(f"🔍 Raw data received: {len(rows)} rows with headers: {len(headers)} columns")
        print(f"📋 Headers: {headers[:5]}...")  # Show first 5 headers for debugging
        
        # Filter empty rows - MATCH FRONTEND LOGIC EXACTLY
        # Frontend: row.some((cell) => cell !== undefined && cell !== null && cell !== "")
        filtered_rows = []
        for row in rows:
            # Check if row has any non-empty cell (matching frontend logic)
            has_content = any(
                cell is not None and 
                cell != "" and 
                str(cell).strip() != "" and
                not (isinstance(cell, float) and pd.isna(cell))
                for cell in row
            )
            if has_content:
                filtered_rows.append(row)
        
        rows = filtered_rows
        print(f"🧹 After filtering empty rows: {len(rows)} rows remaining")
        
        # Extract years for holiday calculation
        years = self._extract_years_from_data(headers, rows)
        self.holidays = get_holidays_for_years(years)
        
        # Clean up headers and data
        headers, rows = self._clean_data(headers, rows)
        
        # Remove specific columns that shouldn't be in output (match frontend exactly)
        headers, rows = self._remove_unwanted_columns(headers, rows)
        
        # Sort data by Request ID and date/time
        rows = self._sort_data_by_request_id(headers, rows)
        
        # Get header indices
        header_indices = self._get_header_indices(headers)
        
        # Add new columns for calculations
        self._add_calculation_columns(headers, header_indices)
        
        # First pass: basic calculations
        first_pass_rows = self._first_pass_calculations(headers, rows, header_indices)
        
        # Second pass: advanced calculations with dependencies
        processed_rows = self._second_pass_calculations(headers, first_pass_rows, header_indices)
        
        final_data = [headers] + processed_rows
        print(f"🎯 Final processed data: {len(final_data)} total rows (including header)")
        print(f"📋 Final headers: {len(headers)} columns")
        print(f"📊 Final data rows: {len(processed_rows)} rows")
        
        return final_data
    
    def _extract_years_from_data(self, headers, rows):
        """Extract years from date columns for holiday calculation"""
        years = []
        
        req_creation_idx = None
        historical_change_idx = None
        
        for i, header in enumerate(headers):
            header_str = str(header).strip()
            if header_str == "Req. Creation Date":
                req_creation_idx = i
            elif header_str == "Historical Status - Change Date":
                historical_change_idx = i
        
        for row in rows:
            # Extract from creation date
            if req_creation_idx is not None and len(row) > req_creation_idx:
                date_val = row[req_creation_idx]
                if date_val and not pd.isna(date_val):
                    year = self._extract_year_from_date(str(date_val))
                    if year:
                        years.append(year)
            
            # Extract from change date  
            if historical_change_idx is not None and len(row) > historical_change_idx:
                date_val = row[historical_change_idx]
                if date_val and not pd.isna(date_val):
                    year = self._extract_year_from_date(str(date_val))
                    if year:
                        years.append(year)
        
        return years
    
    def _extract_year_from_date(self, date_str):
        """Extract year from date string"""
        try:
            if '/' in date_str:
                parts = date_str.split('/')
                if len(parts) == 3:
                    # Try different formats
                    for part in parts:
                        if len(part) == 4 and part.isdigit():
                            return part
            return None
        except:
            return None
    
    def _clean_data(self, headers, rows):
        """Clean headers and remove empty columns - MATCH FRONTEND LOGIC"""
        print(f"🧹 Cleaning data: {len(headers)} headers, {len(rows)} rows")
        
        # Find date column indices
        req_creation_idx = None
        historical_change_idx = None
        
        for i, header in enumerate(headers):
            header_str = str(header).strip()
            if header_str == "Req. Creation Date":
                req_creation_idx = i
            elif header_str == "Historical Status - Change Date":
                historical_change_idx = i
        
        # ENSURE ALL ROWS HAVE SAME LENGTH AS HEADERS (like frontend)
        # Frontend: while (newRow.length < headers.length) newRow.push("");
        for row in rows:
            while len(row) < len(headers):
                row.append("")
        
        print(f"📏 After padding rows to header length: {len(headers)} columns per row")
        
        # Convert date formats in rows (like frontend)
        print(f"🗓️ Converting date formats for columns: req_creation_idx={req_creation_idx}, historical_change_idx={historical_change_idx}")
        date_conversion_count = 0
        for row_idx, row in enumerate(rows):
            if req_creation_idx is not None and len(row) > req_creation_idx:
                cell_value = row[req_creation_idx]
                if cell_value and not pd.isna(cell_value):
                    try:
                        original_value = cell_value
                        converted_value = DateUtils.convert_excel_date(cell_value)
                        row[req_creation_idx] = converted_value
                        if row_idx < 3:  # Debug first few rows
                            print(f"🗓️ Row {row_idx}: Creation date '{original_value}' -> '{converted_value}'")
                        date_conversion_count += 1
                    except Exception as e:
                        print(f"⚠️ Warning: Failed to convert creation date '{cell_value}' at row {row_idx}: {e}")
                        row[req_creation_idx] = ""
                        
            if historical_change_idx is not None and len(row) > historical_change_idx:
                cell_value = row[historical_change_idx]
                if cell_value and not pd.isna(cell_value):
                    try:
                        original_value = cell_value
                        converted_value = DateUtils.convert_excel_date(cell_value)
                        row[historical_change_idx] = converted_value
                        if row_idx < 3:  # Debug first few rows
                            print(f"🗓️ Row {row_idx}: Change date '{original_value}' -> '{converted_value}'")
                        date_conversion_count += 1
                    except Exception as e:
                        print(f"⚠️ Warning: Failed to convert change date '{cell_value}' at row {row_idx}: {e}")
                        row[historical_change_idx] = ""
        
        print(f"🗓️ Converted {date_conversion_count} date values")
        
        # Identify empty columns - MATCH FRONTEND LOGIC EXACTLY
        # Frontend: Only remove columns if BOTH header is empty AND all data is empty
        empty_columns = []
        for col in range(len(headers)):
            # Check if header is empty
            is_header_empty = (not headers[col] or 
                             headers[col] == "" or 
                             str(headers[col]).strip() == "")
            
            # Check if all data in column is empty
            is_column_empty = True
            for row in rows:
                if len(row) > col:
                    cell = row[col]
                    # Match frontend logic: cell !== "" && cell !== undefined && cell !== null
                    if (cell is not None and 
                        cell != "" and 
                        str(cell).strip() != "" and
                        not (isinstance(cell, float) and pd.isna(cell))):
                        is_column_empty = False
                        break
            
            # Only remove if BOTH header AND data are empty (like frontend)
            if is_header_empty and is_column_empty:
                empty_columns.append(col)
                print(f"🗑️ Marking column {col} for removal: header='{headers[col]}', isEmpty={is_column_empty}")
        
        print(f"📊 Empty columns to remove: {empty_columns}")
        
        # Remove empty columns (in reverse order to avoid index shifting)
        for col in reversed(empty_columns):
            print(f"🗑️ Removing column {col}: '{headers[col]}'")
            headers.pop(col)
            for row in rows:
                if len(row) > col:
                    row.pop(col)
        
        print(f"✅ After cleaning: {len(headers)} headers, {len(rows)} rows")
        
        # Sort data by Request ID and change date/time - MATCH FRONTEND LOGIC EXACTLY
        # Temporarily disabled to debug
        # headers, rows = self._sort_data_by_request_id(headers, rows)
        print(f"📊 Sorting temporarily disabled for debugging")
        return headers, rows
    
    def _sort_data_by_request_id(self, headers, rows):
        """Sort data by Request ID and change date/time - MATCH FRONTEND LOGIC EXACTLY"""
        request_id_index = -1
        change_date_index = -1
        change_time_index = -1
        
        # Find column indices
        for i, header in enumerate(headers):
            header_str = str(header).strip()
            if header_str == "Request - ID":
                request_id_index = i
            elif header_str == "Historical Status - Change Date":
                change_date_index = i
            elif header_str == "Historical Status - Change Time":
                change_time_index = i
        
        if request_id_index == -1:
            print("⚠️ Warning: No Request - ID column found, skipping sort")
            return headers, rows
        
        print(f"🔍 Sorting with Request-ID idx={request_id_index}, Change Date idx={change_date_index}, Change Time idx={change_time_index}")
        
        # Group by Request ID
        grouped_data = {}
        for row in rows:
            if len(row) > request_id_index:
                request_id = str(row[request_id_index])
                if request_id not in grouped_data:
                    grouped_data[request_id] = []
                grouped_data[request_id].append(row)
        
        # Sort each group by date/time if available
        for request_id in grouped_data:
            if change_date_index != -1 and change_time_index != -1:
                try:
                    grouped_data[request_id].sort(key=lambda row: self._get_datetime_sort_key(row, change_date_index, change_time_index))
                    print(f"✅ Sorted group {request_id} with {len(grouped_data[request_id])} rows")
                except Exception as e:
                    print(f"⚠️ Error sorting group {request_id}: {e}")
                    # Keep original order if sorting fails
        
        # Flatten the grouped data back into an array
        sorted_rows = []
        for request_id in sorted(grouped_data.keys()):  # Sort by request ID for consistency
            sorted_rows.extend(grouped_data[request_id])
        
        print(f"📊 Sorted {len(sorted_rows)} rows into {len(grouped_data)} groups")
        print(f"🔍 Returning: headers type={type(headers)}, rows type={type(sorted_rows)}")
        return headers, sorted_rows
    
    def _get_datetime_sort_key(self, row, change_date_index, change_time_index):
        """Get sort key for date/time sorting - MATCH FRONTEND LOGIC EXACTLY"""
        try:
            if len(row) <= max(change_date_index, change_time_index):
                return (0, 0, 0, 0, 0, 0)  # Default sort key
                
            date_str = str(row[change_date_index]).strip()
            time_str = str(row[change_time_index]).strip()
            
            # Parse date (expect DD/MM/YYYY or MM/DD/YYYY)
            if '/' in date_str:
                parts = date_str.split('/')
                if len(parts) == 3:
                    # Try DD/MM/YYYY first (more common in input)
                    try:
                        day, month, year = parts
                        year = int(year)
                        month = int(month)
                        day = int(day)
                        
                        # If day > 12, assume DD/MM/YYYY format
                        if day > 12:
                            pass  # day and month are correct
                        # If month > 12, assume MM/DD/YYYY format
                        elif month > 12:
                            day, month = month, day
                        # If both <= 12, we need to guess - use backend format detection logic
                        else:
                            # Use the same detection logic as DateUtils.to_mmddyyyy_from_any
                            # For now, assume DD/MM/YYYY (matches the user's input format)
                            pass
                            
                    except ValueError:
                        return (0, 0, 0, 0, 0, 0)
                else:
                    return (0, 0, 0, 0, 0, 0)
            else:
                return (0, 0, 0, 0, 0, 0)
            
            # Parse time (expect HHMMSS or HH:MM:SS)
            time_str = time_str.replace(':', '').replace(' ', '').zfill(6)
            if len(time_str) >= 6 and time_str.isdigit():
                hours = int(time_str[0:2])
                minutes = int(time_str[2:4])
                seconds = int(time_str[4:6])
            else:
                hours = minutes = seconds = 0
            
            # Return sort tuple: (year, month, day, hours, minutes, seconds)
            return (year, month, day, hours, minutes, seconds)
            
        except (ValueError, IndexError) as e:
            print(f"⚠️ Warning: Failed to parse date/time for sorting: {e}")
            return (0, 0, 0, 0, 0, 0)
    
    def _remove_unwanted_columns(self, headers, rows):
        """Remove columns that appear in source data but not in frontend output"""
        columns_to_remove = [
            "Request - Subject description"
        ]
        
        # Find indices of columns to remove
        remove_indices = []
        for i, header in enumerate(headers):
            header_str = str(header).strip()
            if header_str in columns_to_remove:
                remove_indices.append(i)
                print(f"🗑️ Marking column '{header_str}' for removal (index {i})")
        
        # Remove columns in reverse order to avoid index shifting
        for col_idx in reversed(remove_indices):
            print(f"🗑️ Removing unwanted column {col_idx}: '{headers[col_idx]}'")
            headers.pop(col_idx)
            for row in rows:
                if len(row) > col_idx:
                    row.pop(col_idx)
        
        print(f"✅ After removing unwanted columns: {len(headers)} headers remaining")
        return headers, rows
    
    def _sort_data_by_request_id(self, headers, rows):
        """Sort data by Request ID and then by date/time"""
        request_id_idx = None
        change_date_idx = None
        change_time_idx = None
        
        for i, header in enumerate(headers):
            header_str = str(header).lower()
            if "request - id" in header_str or "request id" in header_str:
                request_id_idx = i
            elif "historical status - change date" in header_str:
                change_date_idx = i
            elif "historical status - change time" in header_str:
                change_time_idx = i
        
        if request_id_idx is None:
            return rows
        
        # Group by Request ID
        grouped_data = {}
        for row in rows:
            if len(row) > request_id_idx:
                request_id = row[request_id_idx]
                if request_id not in grouped_data:
                    grouped_data[request_id] = []
                grouped_data[request_id].append(row)
        
        # Sort each group by date/time
        for request_id in grouped_data:
            grouped_data[request_id].sort(key=lambda row: self._get_sort_key(row, change_date_idx, change_time_idx))
        
        # Flatten back to single list
        return [row for group in grouped_data.values() for row in group]
    
    def _get_sort_key(self, row, change_date_idx, change_time_idx):
        """Generate sort key for date/time sorting"""
        try:
            if change_date_idx is not None and len(row) > change_date_idx:
                date_str = str(row[change_date_idx])
                if '/' in date_str:
                    parts = date_str.split('/')
                    if len(parts) == 3:
                        day, month, year = parts
                        date_key = f"{year}{month.zfill(2)}{day.zfill(2)}"
                        
                        if change_time_idx is not None and len(row) > change_time_idx:
                            time_str = str(row[change_time_idx]).zfill(6)
                            time_key = f"{time_str[:2]}{time_str[2:4]}{time_str[4:6]}"
                            return f"{date_key}{time_key}"
                        
                        return date_key
            return "00000000000000"  # Default sort key
        except:
            return "00000000000000"
    
    def _get_header_indices(self, headers):
        """Get indices of important headers"""
        indices = {}
        
        header_map = {
            'req_creation_date': ['Req. Creation Date'],
            'creation_time': ['Creation Time'],
            'historical_status_from': ['Historical Status - Status From'],
            'request_id': ['Request - ID'],
            'historical_status_to': ['Historical Status - Status To'],
            'req_status_description': ['Req. Status - Description'],
            'historical_change_date': ['Historical Status - Change Date'],
            'historical_change_time': ['Historical Status - Change Time'],
            'priority_description': ['Request - Priority Description'],
            'req_closing_date': ['Req. Closing Date'],
            'req_type_description': ['Req. Type - Description EN']
        }
        
        for key, possible_headers in header_map.items():
            indices[key] = None
            for i, header in enumerate(headers):
                header_str = str(header).strip()
                if any(ph == header_str for ph in possible_headers):
                    indices[key] = i
                    break
        
        return indices
    
    def _add_calculation_columns(self, headers, header_indices):
        """Add new columns for calculations"""
        for field in YELLOW_FIELDS:
            if field not in headers:
                headers.append(field)
                header_indices[field.lower()] = len(headers) - 1
            else:
                header_indices[field.lower()] = headers.index(field)
    
    def _first_pass_calculations(self, headers, rows, header_indices):
        """First pass calculations"""
        print(f"🧮 Starting first pass calculations for {len(rows)} rows")
        processed_rows = []
        
        for i, row in enumerate(rows):
            new_row = row[:]
            # Ensure row has enough elements
            while len(new_row) < len(headers):
                new_row.append("")
            
            # Get values from row
            status_from = str(new_row[header_indices['historical_status_from']] or "").strip()
            status_to = str(new_row[header_indices['historical_status_to']] or "").strip()
            request_id = new_row[header_indices['request_id']]
            priority = str(new_row[header_indices['priority_description']] or "P4 - Low").strip()
            priority_level = priority.split(" ")[0] if priority else "P4"
            
            # Date/time parsing
            creation_date_val = new_row[header_indices['req_creation_date']]
            creation_time = new_row[header_indices['creation_time']]
            change_date_val = new_row[header_indices['historical_change_date']]
            change_time = new_row[header_indices['historical_change_time']]
            req_type_desc = str(new_row[header_indices['req_type_description']] or "").strip()
            
            # Parse date/time with error handling
            try:
                creation_datetime = DateUtils.parse_date_time(creation_date_val, creation_time)
            except Exception as e:
                print(f"⚠️ Warning: Failed to parse creation datetime at row {i}: {e}")
                creation_datetime = None
                
            try:
                change_datetime = DateUtils.parse_date_time(change_date_val, change_time)
            except Exception as e:
                print(f"⚠️ Warning: Failed to parse change datetime at row {i}: {e}")
                change_datetime = None
            
            # ResolSLA calculation
            allowed_statuses_to = [
                "Work in progress", "Forwarded", "Assigned", 
                "Solved", "Suspended", "Pending for IT check", "Awaiting external provider"
            ]
            excluded_statuses_from = [
                "Suspended", "Pending for IT check", 
                "Awaiting external provider"
            ]
            
            is_allowed_to = any(s.lower() == status_to.lower() for s in allowed_statuses_to)
            is_excluded_from = any(s.lower() == status_from.lower() for s in excluded_statuses_from)
            
            new_row[header_indices['resolsla']] = "Yes" if is_allowed_to and not is_excluded_from else " "
            
            # RespSLA calculation
            prev_request_id = rows[i-1][header_indices['request_id']] if i > 0 else None
            new_row[header_indices['respsla']] = "Yes" if i == 0 or request_id != prev_request_id else " "
            
            # ReqCrDtConc calculation - MATCH FRONTEND LOGIC EXACTLY
            # Frontend requires RespSLA == "Yes" AND creationDateTime truthy
            if new_row[header_indices['respsla']] == "Yes" and creation_datetime:
                try:
                    # Use the parsed datetime to get consistent MM/DD/YYYY format
                    if creation_datetime and hasattr(creation_datetime, 'year'):
                        formatted_date = f"{creation_datetime.month:02d}/{creation_datetime.day:02d}/{creation_datetime.year}"
                    else:
                        # Fallback to conversion function
                        formatted_date = DateUtils.to_mmddyyyy_from_any(creation_date_val, creation_datetime)
                    
                    creation_time_str = DateUtils.format_time(creation_time)
                    new_row[header_indices['reqcrdtconc']] = f"{formatted_date} {creation_time_str}"
                except Exception as e:
                    print(f"⚠️ Warning: Failed to format creation date at row {i}: {e}")
                    new_row[header_indices['reqcrdtconc']] = " "
            else:
                new_row[header_indices['reqcrdtconc']] = " "
            
            # EnDtConc calculation
            new_row[header_indices['endtconc']] = DateUtils.format_time(change_time)
            
            # HisChDtTiConc calculation (normalize date)
            if change_datetime:
                try:
                    # Ensure we get the MM/DD/YYYY format consistently
                    if change_datetime and hasattr(change_datetime, 'year'):
                        # Use the parsed datetime to get consistent MM/DD/YYYY format
                        formatted_date = f"{change_datetime.month:02d}/{change_datetime.day:02d}/{change_datetime.year}"
                    else:
                        # Fallback to conversion function
                        formatted_date = DateUtils.to_mmddyyyy_from_any(change_date_val, change_datetime)
                    
                    formatted_time = DateUtils.format_time(change_time)
                    new_row[header_indices['hischdtticonc']] = f"{formatted_date} {formatted_time}"
                except Exception as e:
                    print(f"⚠️ Warning: Failed to format change date at row {i}: {e}")
                    new_row[header_indices['hischdtticonc']] = " "
            else:
                new_row[header_indices['hischdtticonc']] = " "
            
            # SOW values from SLA table
            sla_config = SLA_TABLE.get(priority_level, SLA_TABLE["P4"])
            new_row[header_indices['resolsow']] = sla_config["resolsow"]
            new_row[header_indices['respsow']] = sla_config["respsow"]
            
            processed_rows.append(new_row)
        
        print(f"✅ First pass completed: {len(processed_rows)} rows processed")
        return processed_rows
    
    def _second_pass_calculations(self, headers, rows, header_indices):
        """Second pass calculations with dependencies"""
        print(f"🧮 Starting second pass calculations for {len(rows)} rows")
        processed_rows = []
        last_processed_row = None
        
        for i, row in enumerate(rows):
            new_row = row[:]
            request_id = new_row[header_indices['request_id']]
            prev_row = last_processed_row if i > 0 else None
            next_row = rows[i + 1] if i < len(rows) - 1 else None
            prev_request_id = prev_row[header_indices['request_id']] if prev_row else None
            next_request_id = next_row[header_indices['request_id']] if next_row else None
            status_to = str(new_row[header_indices['historical_status_to']] or "").strip()
            req_status_desc = str(new_row[header_indices['req_status_description']] or "").strip()
            req_type_desc = str(new_row[header_indices['req_type_description']] or "").strip()
            
            # CalcStDt calculation
            calcstdt_condition = (new_row[header_indices['resolsla']] == "Yes" and 
                new_row[header_indices['reqcrdtconc']] and new_row[header_indices['reqcrdtconc']] != " " and
                new_row[header_indices['hischdtticonc']] and new_row[header_indices['hischdtticonc']] != " ")
            
            # Debug for development (remove in production)
            # if request_id in ["A407478L", "A460679L"]:
            #     print(f"🔍 CalcStDt Debug Row {i} ({request_id}): condition = {calcstdt_condition}")
            
            if calcstdt_condition:
                start_date = parse_custom_date(new_row[header_indices['reqcrdtconc']])
                end_date = parse_custom_date(new_row[header_indices['hischdtticonc']])
                
                if start_date and end_date:
                    working_hours = CalculationUtils.calculate_pre_dt(
                        start_date, end_date, WORK_HOURS['start'], WORK_HOURS['end'], self.holidays, i
                    )
                    new_row[header_indices['calcstdt']] = f"{working_hours:.2f}"
                    
                    # Debug (remove in production)
                    # if request_id in ["A407478L", "A460679L"]:
                    #     print(f"   ✅ CalcStDt calculated: {working_hours:.2f}")
                else:
                    new_row[header_indices['calcstdt']] = "0"
            else:
                new_row[header_indices['calcstdt']] = "0"
            
            # RefinedStDt calculation
            calc_st_dt = float(new_row[header_indices['calcstdt']] or 0)
            is_service_request = req_type_desc == "Service Request"
            is_holiday = convert_to_iso_date(new_row[header_indices['historical_change_date']]) in self.holidays
            
            if calc_st_dt < 0 or is_service_request or is_holiday:
                new_row[header_indices['refinedstdt']] = "0"
            else:
                new_row[header_indices['refinedstdt']] = new_row[header_indices['calcstdt']]
            
            # CalcPreDt calculation - MATCH FRONTEND LOGIC EXACTLY
            # Frontend: if (respsla !== "Yes" && requestId === prevRequestId && prevRow?.[headerIndices.hischdtticonc])
            calcpredt_condition = (
                new_row[header_indices['respsla']] != "Yes" and 
                request_id == prev_request_id and 
                prev_row and 
                prev_row[header_indices['hischdtticonc']] and 
                prev_row[header_indices['hischdtticonc']] != " "
            )
            
            # Debug for development (remove in production)
            # if request_id in ["A407478L", "A460679L"]:
            #     print(f"🔍 CalcPreDt Debug Row {i} ({request_id}): condition = {calcpredt_condition}")
            
            if calcpredt_condition:
                start_date = parse_custom_date(prev_row[header_indices['hischdtticonc']])
                end_date = parse_custom_date(new_row[header_indices['hischdtticonc']] or datetime.now().strftime("%m/%d/%Y %H:%M:%S"))
                
                if start_date and end_date:
                    working_hours = CalculationUtils.calculate_pre_dt(
                        start_date, end_date, WORK_HOURS['start'], WORK_HOURS['end'], self.holidays, i
                    )
                    new_row[header_indices['calcpredt']] = f"{working_hours:.2f}"
                    
                    # Debug (remove in production)
                    # if request_id in ["A407478L", "A460679L"]:
                    #     print(f"   ✅ CalcPreDt calculated: {working_hours:.2f}")
                else:
                    new_row[header_indices['calcpredt']] = "0.00"
            else:
                new_row[header_indices['calcpredt']] = "0.00"
            
            # ReqComp calculation
            if status_to in ["Closed", "Discarded"]:
                new_row[header_indices['reqcomp']] = "End"
            elif next_row and request_id != next_request_id:
                new_row[header_indices['reqcomp']] = "Open"
            else:
                new_row[header_indices['reqcomp']] = " "
            
            # RefinedPreDt calculation
            calc_pre_dt = float(new_row[header_indices['calcpredt']] or 0)
            if calc_pre_dt < 0 or is_service_request or is_holiday:
                new_row[header_indices['refinedpredt']] = "0"
            else:
                new_row[header_indices['refinedpredt']] = new_row[header_indices['calcpredt']]
            
            # ElapsedTime calculation - MATCH FRONTEND LOGIC EXACTLY
            # Frontend: if (resolsla === "Yes" && (respsla === " " || respsla === " ")) use refinedpredt, else use refinedstdt
            use_refined_predt = (new_row[header_indices['resolsla']] == "Yes" and 
                                 (new_row[header_indices['respsla']] == " " or new_row[header_indices['respsla']] == " "))
            
            if use_refined_predt:
                elapsed_time = float(new_row[header_indices['refinedpredt']] or 0)
            else:
                elapsed_time = float(new_row[header_indices['refinedstdt']] or 0)
            
            new_row[header_indices['elapsedtime']] = f"{elapsed_time:.2f}"
            
            # Cumilative calculation
            cumulative_hours = 0
            if request_id == prev_request_id and prev_row:
                cumulative_hours = float(prev_row[header_indices['cumilative']] or 0)
            cumulative_hours += float(new_row[header_indices['elapsedtime']] or 0)
            new_row[header_indices['cumilative']] = f"{cumulative_hours:.2f}" if cumulative_hours > 0 else "0.00"
            
            # ResolRem calculation
            if request_id != next_request_id:
                resol_sow = float(new_row[header_indices['resolsow']] or 0)
                cumulative = float(new_row[header_indices['cumilative']] or 0)
                new_row[header_indices['resolrem']] = f"{resol_sow - cumulative:.2f}"
            else:
                new_row[header_indices['resolrem']] = "0"
            
            # RespRem calculation
            if new_row[header_indices['respsla']] == "Yes":
                resp_sow = float(new_row[header_indices['respsow']] or 0)
                calc_st_dt = float(new_row[header_indices['calcstdt']] or 0)
                resp_rem = resp_sow - calc_st_dt
            else:
                resp_rem = float(prev_row[header_indices['resprem']] if prev_row else 0)
            
            new_row[header_indices['resprem']] = f"{resp_rem:.2f}"
            
            # Rollover calculation - FOLLOW FRONTEND
            if request_id == next_request_id:
                new_row[header_indices['rollover']] = "2000 01"
            elif req_status_desc not in ["Closed", "Discarded"]:
                today = datetime.now()
                new_row[header_indices['rollover']] = f"{today.year} {today.month:02d}"
            else:
                try:
                    hist_change = new_row[header_indices['hischdtticonc']]
                    if hist_change and hist_change != " ":
                        date_part = hist_change.split(" ")[0]
                        month, day, year = date_part.split("/")
                        change_date = datetime(int(year), int(month), int(day))
                        new_row[header_indices['rollover']] = f"{change_date.year} {change_date.month:02d}"
                    else:
                        new_row[header_indices['rollover']] = " "
                except Exception:
                    new_row[header_indices['rollover']] = " "
            
            # ReqCrYM calculation - FOLLOW FRONTEND
            current_rollover = new_row[header_indices['rollover']]
            if current_rollover and current_rollover.strip() != "":
                if new_row[header_indices['req_creation_date']]:
                    try:
                        creation_date_str = new_row[header_indices['req_creation_date']]
                        # Frontend parses DD/MM/YYYY here (see frontend block)
                        day, month, year = creation_date_str.split("/")
                        creation_date = datetime(int(year), int(month), int(day))
                        new_row[header_indices['reqcrym']] = f"{creation_date.year} {creation_date.month:02d}"
                    except Exception:
                        new_row[header_indices['reqcrym']] = " "
                else:
                    new_row[header_indices['reqcrym']] = " "
            else:
                new_row[header_indices['reqcrym']] = "9999 12"
            
            # DateRollover and DateReqCrYM (mirror frontend ordering)
            new_row[header_indices['daterollover']] = new_row[header_indices['rollover']]
            new_row[header_indices['datereqcrym']] = new_row[header_indices['reqcrym']]
            
            last_processed_row = new_row
            processed_rows.append(new_row)
        
        print(f"✅ Second pass completed: {len(processed_rows)} rows processed")
        return processed_rows
    
    def _create_excel_output(self, processed_data):
        """Create formatted Excel output"""
        headers = processed_data[0]
        rows = processed_data[1:]
        
        # Create workbook and worksheet
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "ProcessedData"
        
        # Write headers
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                # Format dates and times properly to match frontend output
                if col_idx <= len(headers):
                    header = headers[col_idx - 1]
                    
                    # Handle datetime objects - convert to MM/DD/YYYY format (like frontend)
                    if isinstance(value, datetime):
                        # For date columns, format as MM/DD/YYYY (frontend format)
                        if any(date_field in str(header) for date_field in ['Date', 'Creation Date', 'Change Date']):
                            value = value.strftime("%m/%d/%Y")
                        else:
                            # For other datetime columns, format as MM/DD/YYYY HH:MM:SS
                            value = value.strftime("%m/%d/%Y %H:%M:%S")
                    
                    # Handle string values that might be dates
                    elif any(date_field in str(header) for date_field in ['Date', 'Creation Date', 'Change Date']):
                        if isinstance(value, str) and value.strip():
                            # Convert various date formats to MM/DD/YYYY (frontend format)
                            try:
                                if '/' in value:
                                    parts = value.split('/')
                                    if len(parts) == 3 and all(p.isdigit() for p in parts):
                                        # Detect format based on length and position
                                        if len(parts[2]) == 4:  # YYYY is last
                                            if len(parts[0]) == 2 and len(parts[1]) == 2:
                                                # Could be MM/DD/YYYY or DD/MM/YYYY  
                                                # Frontend typically uses MM/DD/YYYY
                                                month, day, year = parts
                                                value = f"{month}/{day}/{year}"
                                elif '-' in value and len(value) > 10:
                                    # Handle "2020-11-26 00:00:00" format
                                    date_part = value.split(' ')[0]  # Get just the date part
                                    parts = date_part.split('-')
                                    if len(parts) == 3:
                                        year, month, day = parts
                                        value = f"{month}/{day}/{year}"
                            except:
                                pass  # Keep original value if conversion fails
                    
                    # Format time columns
                    elif 'Time' in str(header) and isinstance(value, str):
                        if len(value) == 6 and value.isdigit():
                            value = f"{value[:2]}:{value[2:4]}:{value[4:6]}"
                
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Apply highlighting to new columns
        highlight_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        bold_font = Font(bold=True)
        
        for col_idx, header in enumerate(headers, 1):
            if header in YELLOW_FIELDS:
                # Highlight header
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = highlight_fill
                cell.font = bold_font
                
                # Highlight data cells
                for row_idx in range(2, len(rows) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = highlight_fill
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

# Main processing function for external use
def process_sla_file(file_bytes: bytes, filename: str) -> Tuple[str, bytes]:
    """
    Main entry point for SLA file processing
    
    Args:
        file_bytes: Raw file bytes (CSV or Excel)
        filename: Original filename
        
    Returns:
        Tuple of (output_filename, processed_file_bytes)
    """
    processor = SLAProcessor()
    return processor.process_file_data(file_bytes, filename)

def process_sla_file_enhanced(
    file_bytes: bytes, 
    filename: str, 
    output_format: str = "excel", 
    clean_format: bool = True
) -> Tuple[str, bytes]:
    """
    Enhanced SLA file processing with format options
    
    Args:
        file_bytes: Raw file bytes (CSV or Excel)
        filename: Original filename
        output_format: "excel" or "csv"
        clean_format: If True, truncates long text fields for better visibility
        
    Returns:
        Tuple of (output_filename, processed_file_bytes)
    """
    processor = SLAProcessor()
    return processor.process_file_data_enhanced(file_bytes, filename, output_format, clean_format)

